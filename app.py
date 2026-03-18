# app.py  — versión completa

import os
import re
import csv
import json
import uuid
import zipfile
import sqlite3
import tempfile
import unicodedata

from datetime import datetime
from functools import wraps
from io import BytesIO, StringIO

from flask import (
    Flask, render_template, request, redirect, url_for,
    abort, flash, session, send_file, has_request_context,
    jsonify, after_this_request
)

import qrcode
from werkzeug.security import generate_password_hash, check_password_hash

# ---------- ReportLab (PDF) ----------
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.utils import ImageReader
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

from pypdf import PdfReader, PdfWriter

# ---------- Excel (openpyxl) ----------
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


# =========================================
# APP
# =========================================
app = Flask(__name__)
app.secret_key = "dev-secret"  # cambia en producción


# =========================================
# CONFIG
# =========================================
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "db.sqlite3")
QR_DIR = os.path.join(BASE_DIR, "static", "qrs")
STATIC_DIR = os.path.join(BASE_DIR, "static")
CERT_STATIC = os.path.join(STATIC_DIR, "certs")
FONTS_DIR = os.path.join(CERT_STATIC, "fonts")

os.makedirs(QR_DIR, exist_ok=True)
os.makedirs(CERT_STATIC, exist_ok=True)
os.makedirs(FONTS_DIR, exist_ok=True)

FALLBACK_BASE_URL = (os.environ.get("RENDER_EXTERNAL_URL") or "").rstrip("/")
if not FALLBACK_BASE_URL:
    FALLBACK_BASE_URL = "http://192.168.1.41:5000"

PASSING_GRADE = 60.0
MAX_ZIP_ITEMS = int(os.getenv("MAX_ZIP_ITEMS", "200"))

TEMPLATE_PNG = os.path.join(CERT_STATIC, "plantilla_certificado.png")
TEMPLATE_PNG_INGLES = os.path.join(CERT_STATIC, "plantilla_certificado_ingles.png")
TEMPLATE_PNG_FRANCES = os.path.join(CERT_STATIC, "plantilla_certificado_frances.png")
PLANTILLA_NOTAS_PDF = os.path.join(CERT_STATIC, "plantilla_notas.pdf")
LOGO_WATERMARK = os.path.join(CERT_STATIC, "logo_marca_agua.png")


# =========================================
# CATÁLOGOS
# =========================================
PROGRAMAS = {
    "Inglés": {
        "Principiante": ["PRE A1", "A1", "A1 PLUS", "A1 BASICO"],
        "Pre Intermedia": ["PRE A2", "A2", "A2 Plus"],
        "Intermedia": ["PRE B1", "B1", "B1 PLUS", "PRE B2", "B2", "B2 PLUS"],
        "Avanzada": ["PRE C1", "C1", "C1 PLUS", "PRE C2", "C2", "C2 PLUS"],
        "Curso de vacaciones": ["CVacaciones 1"],
    },
    "Francés": {
        "Principiante": ["PRE A1", "A1", "A1 PLUS", "A1 BASICO"],
        "Pre Intermedia": ["PRE A2", "A2", "A2 Plus"],
        "Intermedia": ["PRE B1", "B1", "B1 PLUS", "PRE B2", "B2", "B2 PLUS"],
        "Avanzada": ["PRE C1", "C1", "C1 PLUS", "PRE C2", "C2", "C2 PLUS"],
        "Curso de vacaciones": ["CVacaciones 1"],
    }
}

# Compatibilidad con plantillas viejas
ETAPAS = PROGRAMAS["Inglés"]

ACTIVIDADES = [
    ("Examen 1", 10, "examen_1"),
    ("Examen 2", 10, "examen_2"),
    ("Lectura", 10, "lectura"),
    ("Escritura", 10, "escritura"),
    ("Vocabulario", 10, "vocabulario"),
    ("Club de Conversación", 10, "club_conversacion"),
    ("Comprensión Auditiva", 10, "comprension_auditiva"),
    ("Examen General", 30, "examen_general"),
]
TOTAL_PUNTOS = 100

ETAPA_ALIASES = {
    "preintermedio": "Pre Intermedia",
    "preintermedia": "Pre Intermedia",
    "preintermediaa": "Pre Intermedia",
    "preintermediate": "Pre Intermedia",
}

MESES_ES = [
    "enero", "febrero", "marzo", "abril", "mayo", "junio",
    "julio", "agosto", "septiembre", "octubre", "noviembre", "diciembre"
]


# =========================================
# HELPERS
# =========================================
def strip_accents_py(s: str) -> str:
    if not s:
        return ""
    return "".join(
        ch for ch in unicodedata.normalize("NFD", s)
        if unicodedata.category(ch) != "Mn"
    )


def _norm_key(s: str) -> str:
    return strip_accents_py(s or "").lower().replace(" ", "").replace("_", "").replace("-", "")


def get_programa(programa: str) -> str:
    programa = (programa or "").strip()
    if programa in PROGRAMAS:
        return programa
    if _norm_key(programa) in ("frances", "francais"):
        return "Francés"
    return "Inglés"


def get_etapas(programa: str) -> dict:
    return PROGRAMAS.get(get_programa(programa), PROGRAMAS["Inglés"])


def canonical_etapa(programa: str, etapa: str) -> str:
    n = _norm_key(etapa)
    etapas = get_etapas(programa)

    for k in etapas.keys():
        if _norm_key(k) == n:
            return k

    if n in ETAPA_ALIASES:
        return ETAPA_ALIASES[n]

    return etapa


def canonical_nivel(programa: str, etapa_canon: str, nivel: str) -> str:
    n = _norm_key(nivel)
    opciones = get_etapas(programa).get(etapa_canon, []) or []
    for opt in opciones:
        if _norm_key(opt) == n:
            return opt
    return nivel


def etapa_keys(programa: str, etapa_canon: str) -> list[str]:
    keys = {_norm_key(etapa_canon)}
    for alias_norm, canon in ETAPA_ALIASES.items():
        if canon == etapa_canon:
            keys.add(alias_norm)
    return sorted(keys)


def _fecha_es(dt: datetime) -> str:
    try:
        return f"{dt.day} de {MESES_ES[dt.month - 1].capitalize()} de {dt.year}"
    except Exception:
        return dt.strftime("%Y-%m-%d")


def conn():
    c = sqlite3.connect(DB_PATH)
    c.row_factory = sqlite3.Row
    c.create_function("strip_accents", 1, strip_accents_py)
    c.create_function("norm_key", 1, _norm_key)
    return c


def _ensure_paths():
    os.makedirs(QR_DIR, exist_ok=True)
    os.makedirs(CERT_STATIC, exist_ok=True)
    os.makedirs(FONTS_DIR, exist_ok=True)


def _has_column(colname: str) -> bool:
    with conn() as c:
        cols = [r[1] for r in c.execute("PRAGMA table_info(estudiantes)").fetchall()]
    return colname in cols


def _get_notas_value(row: sqlite3.Row) -> str:
    return (
        row["notas"] if "notas" in row.keys() and row["notas"]
        else (row["columna"] if "columna" in row.keys() else "")
    )


def _load_detalle_notas(row: sqlite3.Row) -> dict:
    try:
        raw = row["detalle_notas"]
    except Exception:
        raw = None
    if not raw:
        return {}
    try:
        return json.loads(raw) if isinstance(raw, str) else {}
    except Exception:
        return {}


def _register_cert_fonts() -> None:
    def _try_register(alias: str, filenames: list[str], folder: str):
        if not folder or not os.path.isdir(folder):
            return
        lower = {f.lower(): f for f in os.listdir(folder)}
        for cand in filenames:
            real = lower.get(cand.lower())
            if real:
                try:
                    pdfmetrics.registerFont(TTFont(alias, os.path.join(folder, real)))
                    return
                except Exception:
                    pass

    _try_register("GreatVibes", ["GreatVibes-Regular.ttf"], FONTS_DIR)
    _try_register(
        "Playfair-Bold",
        ["PlayfairDisplay-Bold.ttf", "PlayfairDisplay-Italic-VariableFont_wght.ttf"],
        FONTS_DIR
    )
    _try_register("Arial", ["ARIAL.TTF", "Arial.ttf"], FONTS_DIR)
    _try_register("Arial-Bold", ["ARIALBD 1.TTF", "ARIALBD.TTF", "arialbd.ttf"], FONTS_DIR)

    try:
        win = r"C:\Windows\Fonts"
        _try_register("Arial", ["arial.ttf", "ARIAL.TTF"], win)
        _try_register("Arial-Bold", ["arialbd.ttf", "ARIALBD.TTF"], win)
    except Exception:
        pass


def _pick_font(name: str, fallback_bold: bool = False) -> str:
    regs = set(pdfmetrics.getRegisteredFontNames())
    if name in regs:
        return name
    if fallback_bold:
        if "Arial-Bold" in regs:
            return "Arial-Bold"
        return "Helvetica-Bold"
    if "Arial" in regs:
        return "Arial"
    return "Helvetica"


def _safe_filename(name: str) -> str:
    name = (name or "").strip()
    name = re.sub(r"[^a-zA-Z0-9._-]+", "_", name)
    name = name.strip("._-") or "certificado"
    return name[:80]


def _to_datetime_local_str(db_str: str) -> str:
    if not db_str:
        return ""
    try:
        dt = datetime.strptime(db_str, "%Y-%m-%d %H:%M:%S")
        return dt.strftime("%Y-%m-%dT%H:%M")
    except Exception:
        pass
    try:
        dt = datetime.fromisoformat(db_str)
        return dt.strftime("%Y-%m-%dT%H:%M")
    except Exception:
        return ""


def _parse_datetime_local(s: str):
    try:
        return datetime.strptime(s, "%Y-%m-%dT%H:%M")
    except Exception:
        return None


app.jinja_env.filters["dtlocal"] = _to_datetime_local_str


# =========================================
# CONTEXT PROCESSOR
# =========================================
@app.context_processor
def inject_user():
    return {
        "current_user": session.get("user"),
        "PASSING_GRADE": PASSING_GRADE,
        "ETAPAS": ETAPAS,
        "PROGRAMAS": PROGRAMAS,
        "YEAR": datetime.now().year,
    }


# =========================================
# DB INIT
# =========================================
def init_db():
    with conn() as c:
        c.execute("""
        CREATE TABLE IF NOT EXISTS estudiantes(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          token TEXT UNIQUE NOT NULL,
          nombre TEXT NOT NULL,
          curso TEXT NOT NULL,
          nota REAL NOT NULL,
          estado TEXT CHECK(estado IN ('Aprobado','Reprobado')) NOT NULL,
          creado_en TEXT NOT NULL
        )""")

        try:
            c.execute("ALTER TABLE estudiantes ADD COLUMN notas TEXT")
        except Exception:
            pass

        try:
            c.execute("ALTER TABLE estudiantes ADD COLUMN programa TEXT")
        except Exception:
            pass

        try:
            c.execute("UPDATE estudiantes SET programa='Inglés' WHERE programa IS NULL OR programa='' OR programa='AEA'")
        except Exception:
            pass

        try:
            c.execute("ALTER TABLE estudiantes ADD COLUMN detalle_notas TEXT")
        except Exception:
            pass


def init_users():
    with conn() as c:
        c.execute("""
        CREATE TABLE IF NOT EXISTS usuarios(
          id INTEGER PRIMARY KEY AUTOINCREMENT,
          username TEXT UNIQUE NOT NULL,
          password_hash TEXT NOT NULL,
          rol TEXT NOT NULL,
          creado_en TEXT NOT NULL
        )""")

        try:
            c.execute("ALTER TABLE usuarios ADD COLUMN rol TEXT NOT NULL DEFAULT 'admin'")
        except Exception:
            pass

        existentes = c.execute("SELECT username FROM usuarios").fetchall()
        existentes = {u["username"] for u in existentes}

        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        if "rodasestuardo146@gmail.com" not in existentes:
            c.execute(
                """INSERT INTO usuarios(username,password_hash,rol,creado_en)
                   VALUES(?,?,?,?)""",
                (
                    "rodasestuardo146@gmail.com",
                    generate_password_hash("Admin123!@"),
                    "admin",
                    now
                )
            )
            print("✔ Admin fuerte creado")

        if "SubAdminAEA" not in existentes:
            c.execute(
                """INSERT INTO usuarios(username,password_hash,rol,creado_en)
                   VALUES(?,?,?,?)""",
                (
                    "SubAdminAEA",
                    generate_password_hash("SubAdminAEA@"),
                    "subadmin",
                    now
                )
            )
            print("✔ Subadmin creado")


def seed_if_empty():
    with conn() as c:
        n = c.execute("SELECT COUNT(*) FROM estudiantes").fetchone()[0]
        if n == 0:
            token = uuid.uuid4().hex
            estado = "Aprobado" if 85 >= PASSING_GRADE else "Reprobado"
            c.execute("""
                INSERT INTO estudiantes(token,nombre,curso,nota,estado,creado_en,programa)
                VALUES(?,?,?,?,?,?,?)
            """, (
                token, "Alumno Demo", "Intermedia - B1", 85, estado,
                datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "Inglés"
            ))
            build_qr(token)
            print(f"➡ Alumno demo creado. Token: {token}")


with app.app_context():
    try:
        init_db()
        init_users()
        seed_if_empty()
    except Exception as e:
        print("INIT ERROR:", repr(e))


# =========================================
# QR
# =========================================
def build_qr(token: str):
    path = url_for("ver_cert", token=token, _external=False)

    if has_request_context():
        base = request.host_url.rstrip("/")
    else:
        base = (os.environ.get("RENDER_EXTERNAL_URL") or FALLBACK_BASE_URL or "").rstrip("/")

    if not base:
        base = "http://127.0.0.1:5000"

    url = f"{base}{path}"

    os.makedirs(QR_DIR, exist_ok=True)
    img = qrcode.make(url)
    img.save(os.path.join(QR_DIR, f"{token}.png"))


# =========================================
# AUTH
# =========================================
def login_required(view):
    @wraps(view)
    def wrapper(*args, **kwargs):
        user = session.get("user")
        if not user:
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapper


def role_required(*roles):
    def decorator(view):
        @wraps(view)
        def wrapper(*args, **kwargs):
            user = session.get("user")

            if not isinstance(user, dict):
                session.clear()
                flash("Sesión inválida. Inicia sesión nuevamente.", "error")
                return redirect(url_for("login"))

            if user.get("rol") not in roles:
                flash("permiso_denegado", "modal")
                return redirect(url_for("admin"))

            return view(*args, **kwargs)
        return wrapper
    return decorator


# =========================================
# RUTAS PÚBLICAS
# =========================================
@app.route("/")
def index():
    return redirect(url_for("validar"))


@app.route("/validar", methods=["GET", "POST"])
def validar():
    if request.method == "POST":
        nombre_trigger = (request.form.get("nombre") or "").strip()
        trigger_norm = strip_accents_py(nombre_trigger).lower()

        ADMIN_TRIGGERS = {
            "iniciar como administrador",
            "iniciar como admin",
            "admin",
            "administrador"
        }
        if trigger_norm in ADMIN_TRIGGERS:
            return redirect(url_for("login", next=url_for("admin")))

        programa = get_programa(request.form.get("programa") or "Inglés")
        etapa = (request.form.get("etapa") or "").strip()
        nivel = (request.form.get("nivel") or "").strip()

        etapa = canonical_etapa(programa, etapa)
        nivel = canonical_nivel(programa, etapa, nivel)

        if not etapa or not nivel:
            flash("Selecciona etapa y nivel.", "error")
            return redirect(url_for("validar"))

        nombre = nombre_trigger
        if not nombre:
            flash("Ingresa el nombre del estudiante.", "error")
            return redirect(url_for("validar"))

        nombre_q = f"%{strip_accents_py(nombre).lower()}%"

        with conn() as cdb:
            rows = cdb.execute(
                """
                SELECT id, token, nombre, curso, nota, estado, creado_en, programa
                FROM estudiantes
                WHERE lower(strip_accents(nombre)) LIKE ?
                ORDER BY nombre ASC, curso ASC
                """,
                (nombre_q,)
            ).fetchall()

        etapa_key = _norm_key(etapa)
        nivel_key = _norm_key(nivel)
        programa_key = _norm_key(programa)

        filas = []
        for r in rows:
            curso_db = (r["curso"] or "").strip()
            programa_db = get_programa(r["programa"] or "Inglés")

            etapa_db, nivel_db = "", ""
            if " - " in curso_db:
                etapa_db, nivel_db = curso_db.split(" - ", 1)
            else:
                etapa_db = curso_db
                nivel_db = ""

            etapa_db = canonical_etapa(programa_db, etapa_db.strip())
            nivel_db = canonical_nivel(programa_db, etapa_db, (nivel_db or "").strip())

            if (
                _norm_key(programa_db) == programa_key and
                _norm_key(etapa_db) == etapa_key and
                _norm_key(nivel_db) == nivel_key
            ):
                filas.append(r)

        if not filas:
            flash("No se encontró un estudiante con esos datos.", "error")
            return redirect(url_for("validar"))

        if len(filas) == 1:
            return redirect(url_for("ver_cert", token=filas[0]["token"]))

        resultados = [{
            "id": r["id"],
            "token": r["token"],
            "nombre": r["nombre"],
            "curso": r["curso"],
            "programa": r["programa"],
            "cert_url": url_for("ver_cert", token=r["token"], _external=True),
            "pdf_url": url_for("cert_pdf", token=r["token"])
        } for r in filas]

        return render_template(
            "validar.html",
            resultados=resultados,
            q_nombre=nombre,
            q_programa=programa,
            q_etapa=etapa,
            q_nivel=nivel
        )

    return render_template("validar.html")


# =========================================
# PDF – CERTIFICADO
# =========================================
@app.route("/cert/<token>", endpoint="ver_cert")
def certificate_view(token):
    with conn() as c:
        row = c.execute("SELECT * FROM estudiantes WHERE token=?", (token,)).fetchone()
    if not row:
        abort(404)

    e = dict(row)
    e["notas"] = e.get("notas") or e.get("columna") or ""
    e["programa"] = get_programa(e.get("programa") or "Inglés")

    etapa, nivel = "", ""
    curso_val = (e.get("curso") or "").strip()
    if " - " in curso_val:
        etapa, nivel = curso_val.split(" - ", 1)

    png = os.path.join(QR_DIR, f"{token}.png")
    if not os.path.exists(png):
        build_qr(token)

    return render_template(
        "certificado.html",
        est=e,
        etapa=etapa,
        nivel=nivel,
        programa=e["programa"],
        qr_url=f"/static/qrs/{token}.png",
    )


@app.route("/cert/<token>/pdf", endpoint="cert_pdf")
def cert_pdf(token):
    try:
        pdf_bytes = _build_pdf_bytes_from_token(token)
    except Exception as ex:
        abort(500, description=f"No se pudo generar el diploma: {repr(ex)}")

    with conn() as c:
        r = c.execute("SELECT nombre FROM estudiantes WHERE token=?", (token,)).fetchone()
    nombre = (r["nombre"] if r else "certificado") or "certificado"

    buf = BytesIO(pdf_bytes)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=False,
        download_name=f"certificado_{nombre.replace(' ', '_')}.pdf",
        mimetype="application/pdf"
    )


@app.route("/cert/<token>/notas.pdf")
def notas_pdf(token):
    with conn() as c:
        e = c.execute("SELECT * FROM estudiantes WHERE token=?", (token,)).fetchone()
    if not e:
        abort(404)

    detalle = _load_detalle_notas(e)

    nombre = (e["nombre"] or "").strip()
    curso = (e["curso"] or "").strip()
    programa = get_programa(e["programa"] or "Inglés")

    etapa, nivel = "", ""
    if " - " in curso:
        etapa, nivel = curso.split(" - ", 1)
    else:
        etapa = curso

    if detalle:
        total = 0.0
        for _, _, key in ACTIVIDADES:
            try:
                total += float(detalle.get(key, 0) or 0)
            except Exception:
                pass
        try:
            nota_final = float(detalle.get("punteo_final", total))
        except Exception:
            nota_final = total
    else:
        nota_final = float(e["nota"] or 0)

    estado = "APROBADO" if nota_final >= PASSING_GRADE else "REPROBADO"

    try:
        creado = datetime.fromisoformat(e["creado_en"])
        fecha_txt = _fecha_es(creado)
    except Exception:
        fecha_txt = e["creado_en"]

    if detalle:
        idioma = detalle.get("idioma") or ("Francés" if programa == "Francés" else "Inglés Americano")
        nivel_cefr = detalle.get("nivel_cefr") or nivel
    else:
        idioma = "Francés" if programa == "Francés" else "Inglés Americano"
        nivel_cefr = nivel

    if not os.path.exists(PLANTILLA_NOTAS_PDF):
        abort(500, description="No existe static/certs/plantilla_notas.pdf")

    reader = PdfReader(PLANTILLA_NOTAS_PDF)
    base_page = reader.pages[0]

    W = float(base_page.mediabox.width)
    H = float(base_page.mediabox.height)

    overlay_buf = BytesIO()
    cpdf = canvas.Canvas(overlay_buf, pagesize=(W, H))

    X_NOMBRE = 150
    Y_NOMBRE = H - 215

    X_NIVEL = 125
    Y_NIVEL = H - 242

    X_IDIOMA = W - 200
    Y_IDIOMA = H - 220

    X_FECHA = W - 200
    Y_FECHA = H - 242

    X_OBT = 460
    Y_FIRST_ROW = H - 330
    ROW_H = 27

    cpdf.setFillColor(colors.black)

    cpdf.setFont("Helvetica-Bold", 11)
    cpdf.drawString(X_NOMBRE, Y_NOMBRE, nombre)

    cpdf.setFont("Helvetica", 11)
    cpdf.drawString(X_IDIOMA, Y_IDIOMA, idioma)

    cpdf.setFont("Helvetica", 11)
    cpdf.drawString(X_NIVEL, Y_NIVEL, nivel_cefr)

    cpdf.setFont("Helvetica", 11)
    cpdf.drawString(X_FECHA, Y_FECHA, fecha_txt)

    if detalle:
        cpdf.setFont("Helvetica", 11)
        for i, (_, _, key) in enumerate(ACTIVIDADES):
            v = detalle.get(key, "")
            if v == "" or v is None:
                continue
            y = Y_FIRST_ROW - (i * ROW_H)
            cpdf.drawCentredString(X_OBT, y, str(v))

    y_pf = Y_FIRST_ROW - (len(ACTIVIDADES) * ROW_H)
    cpdf.setFont("Helvetica-Bold", 11)
    cpdf.drawCentredString(X_OBT, y_pf, f"{nota_final:.0f}")

    y_res = y_pf - ROW_H
    cpdf.setFont("Helvetica-Bold", 12)
    X_RESULTADO = X_OBT - 60
    cpdf.drawCentredString(X_RESULTADO, y_res, estado)

    cpdf.showPage()
    cpdf.save()
    overlay_buf.seek(0)

    overlay_reader = PdfReader(overlay_buf)
    overlay_page = overlay_reader.pages[0]

    base_page.merge_page(overlay_page)

    out = PdfWriter()
    out.add_page(base_page)

    out_buf = BytesIO()
    out.write(out_buf)
    out_buf.seek(0)

    return send_file(
        out_buf,
        as_attachment=False,
        download_name=f"notas_{nombre.replace(' ', '_')}.pdf",
        mimetype="application/pdf"
    )


# =========================================
# LOGIN / LOGOUT
# =========================================
@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        u = (request.form.get("username") or "").strip()
        p = (request.form.get("password") or "")

        if not u or not p:
            flash("Ingresa usuario y contraseña.", "error")
            return redirect(url_for("login"))

        with conn() as c:
            row = c.execute("SELECT * FROM usuarios WHERE username=?", (u,)).fetchone()

        if row and check_password_hash(row["password_hash"], p):
            session["user"] = {
                "username": row["username"],
                "rol": row["rol"]
            }
            flash("Bienvenido.", "ok")
            return redirect(request.args.get("next") or url_for("admin"))

        flash("Usuario o contraseña incorrectos.", "error")
        return redirect(url_for("login"))

    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    flash("Sesión cerrada.", "ok")
    return redirect(url_for("validar"))


# =========================================
# ADMIN
# =========================================
@app.route("/admin", methods=["GET", "POST"])
@login_required
@role_required("admin", "subadmin")
def admin():
    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        programa = get_programa(request.form.get("programa") or "Inglés")
        etapa = (request.form.get("etapa") or "").strip()
        nivel = (request.form.get("nivel") or "").strip()

        etapa = canonical_etapa(programa, etapa)
        nivel = canonical_nivel(programa, etapa, nivel)

        nota_s = (request.form.get("nota") or "").strip()
        notas_url = (request.form.get("notas") or "").strip()

        if not (nombre and etapa and nivel and nota_s):
            flash("Completa nombre, programa, etapa, nivel y nota.", "error")
            return redirect(url_for("admin"))

        try:
            nota = float(nota_s)
        except Exception:
            flash("La nota debe ser numérica.", "error")
            return redirect(url_for("admin"))

        if not (0 <= nota <= 100):
            flash("La nota debe estar entre 0 y 100.", "error")
            return redirect(url_for("admin"))

        estado = "Aprobado" if nota >= PASSING_GRADE else "Reprobado"
        curso_text = f"{etapa} - {nivel}"
        token = uuid.uuid4().hex
        creado = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        with conn() as c:
            c.execute("""
                INSERT INTO estudiantes(token,nombre,curso,nota,estado,creado_en,notas,programa)
                VALUES(?,?,?,?,?,?,?,?)
            """, (token, nombre, curso_text, nota, estado, creado, notas_url, programa))

        build_qr(token)
        flash("Estudiante creado y QR generado.", "ok")
        return redirect(url_for("admin"))

    q = (request.args.get("q") or "").strip()
    programa_filtro = get_programa(request.args.get("programa") or "Inglés") if (request.args.get("programa") or "").strip() else ""
    etapa_filtro_raw = (request.args.get("etapa") or "").strip()
    nivel_filtro_raw = (request.args.get("nivel") or "").strip()
    estado_filtro = (request.args.get("estado") or "").strip()

    nota_min_s = (request.args.get("nota_min") or "").strip()
    nota_max_s = (request.args.get("nota_max") or "").strip()

    fecha_desde = (request.args.get("desde") or "").strip()
    fecha_hasta = (request.args.get("hasta") or "").strip()

    sort = (request.args.get("sort") or "creado_en").strip()
    order = (request.args.get("order") or "desc").lower()
    page = int(request.args.get("page", 1) or 1)
    per_page = int(request.args.get("per_page", 10) or 10)

    allowed_sorts = {"id", "token", "nombre", "curso", "nota", "estado", "creado_en", "programa"}
    if sort not in allowed_sorts:
        sort = "creado_en"
    if order not in {"asc", "desc"}:
        order = "desc"
    if per_page not in (10, 25, 50, 100):
        per_page = 10
    if page < 1:
        page = 1

    etapa_filtro = canonical_etapa(programa_filtro or "Inglés", etapa_filtro_raw) if etapa_filtro_raw else ""
    nivel_filtro = nivel_filtro_raw.strip()
    if etapa_filtro and nivel_filtro:
        nivel_filtro = canonical_nivel(programa_filtro or "Inglés", etapa_filtro, nivel_filtro)

    niveles_filtro = get_etapas(programa_filtro).get(etapa_filtro, []) if (programa_filtro and etapa_filtro) else []

    where_parts = ["1=1"]
    params = []

    etapa_expr = "trim(CASE WHEN instr(curso,'-')>0 THEN substr(curso,1,instr(curso,'-')-1) ELSE curso END)"
    nivel_expr = "trim(CASE WHEN instr(curso,'-')>0 THEN substr(curso,instr(curso,'-')+1) ELSE '' END)"

    if q:
        q_norm = f"%{strip_accents_py(q).lower()}%"
        where_parts.append("(lower(strip_accents(nombre)) LIKE ? OR lower(strip_accents(curso)) LIKE ? OR token LIKE ?)")
        params.extend([q_norm, q_norm, f"%{q}%"])

    if programa_filtro:
        where_parts.append("norm_key(programa) = ?")
        params.append(_norm_key(programa_filtro))

    if etapa_filtro:
        keys = etapa_keys(programa_filtro or "Inglés", etapa_filtro)
        where_parts.append(f"norm_key({etapa_expr}) IN ({','.join(['?'] * len(keys))})")
        params.extend(keys)

    if nivel_filtro:
        where_parts.append(f"norm_key({nivel_expr}) = ?")
        params.append(_norm_key(nivel_filtro))

    if estado_filtro and estado_filtro.lower() != "todos":
        if estado_filtro not in ("Aprobado", "Reprobado"):
            estado_filtro = "Aprobado" if estado_filtro.lower().startswith("apro") else "Reprobado"
        where_parts.append("estado = ?")
        params.append(estado_filtro)

    def _to_float(s):
        try:
            return float(s)
        except Exception:
            return None

    nmin = _to_float(nota_min_s) if nota_min_s != "" else None
    nmax = _to_float(nota_max_s) if nota_max_s != "" else None

    if nmin is not None:
        where_parts.append("nota >= ?")
        params.append(nmin)
    if nmax is not None:
        where_parts.append("nota <= ?")
        params.append(nmax)

    if fecha_desde:
        where_parts.append("date(creado_en) >= date(?)")
        params.append(fecha_desde)
    if fecha_hasta:
        where_parts.append("date(creado_en) <= date(?)")
        params.append(fecha_hasta)

    where = " AND ".join(where_parts)

    with conn() as cdb:
        total = cdb.execute(f"SELECT COUNT(*) FROM estudiantes WHERE {where}", params).fetchone()[0]
        total_pages = max(1, (total + per_page - 1) // per_page)
        if page > total_pages:
            page = total_pages
        offset = (page - 1) * per_page

        rows = cdb.execute(
            f"SELECT * FROM estudiantes WHERE {where} "
            f"ORDER BY {sort} {order.upper()} LIMIT ? OFFSET ?",
            (*params, per_page, offset),
        ).fetchall()

    listado = [{
        "id": e["id"],
        "token": e["token"],
        "nombre": e["nombre"],
        "curso": e["curso"],
        "programa": e["programa"],
        "nota": e["nota"],
        "estado": e["estado"],
        "creado_en": e["creado_en"],
        "qr_url": f"/static/qrs/{e['token']}.png",
        "cert_url": url_for("ver_cert", token=e["token"], _external=True),
    } for e in rows]

    pages = list(range(max(1, page - 2), min(total_pages, page + 2) + 1))

    return render_template(
        "admin.html",
        estudiantes=listado,
        q=q,

        programa_filtro=programa_filtro,
        etapa_filtro=etapa_filtro,
        nivel_filtro=nivel_filtro,
        niveles_filtro=niveles_filtro,
        estado_filtro=estado_filtro,
        nota_min=nota_min_s,
        nota_max=nota_max_s,

        desde=fecha_desde,
        hasta=fecha_hasta,

        sort=sort,
        order=order,
        page=page,
        per_page=per_page,
        total=total,
        total_pages=total_pages,
        pages=pages,
    )


@app.route("/admin/ids", methods=["GET"])
@login_required
@role_required("admin", "subadmin")
def admin_ids():
    q = (request.args.get("q") or "").strip()
    programa_filtro = get_programa(request.args.get("programa") or "Inglés") if (request.args.get("programa") or "").strip() else ""
    etapa_filtro_raw = (request.args.get("etapa") or "").strip()
    nivel_filtro_raw = (request.args.get("nivel") or "").strip()
    estado_filtro = (request.args.get("estado") or "").strip()

    nota_min_s = (request.args.get("nota_min") or "").strip()
    nota_max_s = (request.args.get("nota_max") or "").strip()
    fecha_desde = (request.args.get("desde") or "").strip()
    fecha_hasta = (request.args.get("hasta") or "").strip()

    etapa_filtro = canonical_etapa(programa_filtro or "Inglés", etapa_filtro_raw) if etapa_filtro_raw else ""
    nivel_filtro = nivel_filtro_raw.strip()
    if etapa_filtro and nivel_filtro:
        nivel_filtro = canonical_nivel(programa_filtro or "Inglés", etapa_filtro, nivel_filtro)

    where_parts = ["1=1"]
    params = []

    etapa_expr = "trim(CASE WHEN instr(curso,'-')>0 THEN substr(curso,1,instr(curso,'-')-1) ELSE curso END)"
    nivel_expr = "trim(CASE WHEN instr(curso,'-')>0 THEN substr(curso,instr(curso,'-')+1) ELSE '' END)"

    if q:
        q_norm = f"%{strip_accents_py(q).lower()}%"
        where_parts.append("(lower(strip_accents(nombre)) LIKE ? OR lower(strip_accents(curso)) LIKE ? OR token LIKE ?)")
        params.extend([q_norm, q_norm, f"%{q}%"])

    if programa_filtro:
        where_parts.append("norm_key(programa) = ?")
        params.append(_norm_key(programa_filtro))

    if etapa_filtro:
        keys = etapa_keys(programa_filtro or "Inglés", etapa_filtro)
        where_parts.append(f"norm_key({etapa_expr}) IN ({','.join(['?'] * len(keys))})")
        params.extend(keys)

    if nivel_filtro:
        where_parts.append(f"norm_key({nivel_expr}) = ?")
        params.append(_norm_key(nivel_filtro))

    if estado_filtro and estado_filtro.lower() != "todos":
        if estado_filtro not in ("Aprobado", "Reprobado"):
            estado_filtro = "Aprobado" if estado_filtro.lower().startswith("apro") else "Reprobado"
        where_parts.append("estado = ?")
        params.append(estado_filtro)

    def _to_float(s):
        try:
            return float(s)
        except Exception:
            return None

    nmin = _to_float(nota_min_s) if nota_min_s != "" else None
    nmax = _to_float(nota_max_s) if nota_max_s != "" else None

    if nmin is not None:
        where_parts.append("nota >= ?")
        params.append(nmin)
    if nmax is not None:
        where_parts.append("nota <= ?")
        params.append(nmax)

    if fecha_desde:
        where_parts.append("date(creado_en) >= date(?)")
        params.append(fecha_desde)
    if fecha_hasta:
        where_parts.append("date(creado_en) <= date(?)")
        params.append(fecha_hasta)

    where = " AND ".join(where_parts)

    with conn() as c:
        ids = [str(r["id"]) for r in c.execute(f"SELECT id FROM estudiantes WHERE {where}", params).fetchall()]

    return jsonify({"ids": ids, "count": len(ids)})


@app.route("/admin/editar/<int:id>", methods=["GET", "POST"])
@login_required
@role_required("admin", "subadmin")
def editar_estudiante(id):
    with conn() as c:
        e = c.execute("SELECT * FROM estudiantes WHERE id=?", (id,)).fetchone()
    if not e:
        abort(404)

    if request.method == "POST":
        nombre = (request.form.get("nombre") or "").strip()
        programa = get_programa(request.form.get("programa") or "Inglés")
        etapa = (request.form.get("etapa") or "").strip()
        nivel = (request.form.get("nivel") or "").strip()

        etapa = canonical_etapa(programa, etapa)
        nivel = canonical_nivel(programa, etapa, nivel)

        nota_s = (request.form.get("nota") or "").strip()
        notas_url = (request.form.get("notas") or "").strip()

        if not (nombre and etapa and nivel and nota_s):
            flash("Completa nombre, programa, etapa, nivel y nota.", "error")
            return redirect(url_for("editar_estudiante", id=id))

        try:
            nota = float(nota_s)
        except Exception:
            flash("La nota debe ser numérica.", "error")
            return redirect(url_for("editar_estudiante", id=id))

        if nota < 0 or nota > 100:
            flash("La nota debe estar entre 0 y 100.", "error")
            return redirect(url_for("editar_estudiante", id=id))

        estado = "Aprobado" if nota >= PASSING_GRADE else "Reprobado"
        curso_text = f"{etapa} - {nivel}"

        with conn() as c:
            c.execute("""
                UPDATE estudiantes
                SET nombre=?, curso=?, nota=?, estado=?, programa=?, notas=?
                WHERE id=?
            """, (nombre, curso_text, nota, estado, programa, notas_url, id))

        flash("Estudiante actualizado.", "ok")
        return redirect(url_for("admin"))

    etapa_val, nivel_val = "", ""
    if " - " in e["curso"]:
        etapa_val, nivel_val = e["curso"].split(" - ", 1)

    programa_val = get_programa(e["programa"] or "Inglés")

    return render_template(
        "editar.html",
        est=e,
        etapa_val=etapa_val,
        nivel_val=nivel_val,
        programa_val=programa_val
    )


@app.route("/admin/eliminar/<int:id>", methods=["POST"])
@login_required
@role_required("admin")
def eliminar_estudiante(id):
    with conn() as c:
        row = c.execute("SELECT token FROM estudiantes WHERE id=?", (id,)).fetchone()
        c.execute("DELETE FROM estudiantes WHERE id=?", (id,))
    if row:
        png = os.path.join(QR_DIR, f"{row['token']}.png")
        if os.path.exists(png):
            try:
                os.remove(png)
            except Exception:
                pass
    flash("Estudiante eliminado.", "ok")
    return redirect(url_for("admin"))


@app.route("/admin/regen/<token>", methods=["POST"])
@login_required
@role_required("admin")
def regenerar_qr(token):
    build_qr(token)
    flash("QR re-generado.", "ok")
    return redirect(url_for("admin"))


# =========================================
# EXPORTAR CSV / EXCEL
# =========================================
@app.route("/admin/export/csv")
@login_required
@role_required("admin")
def export_csv():
    with conn() as c:
        rows = c.execute("""
            SELECT id, token, nombre, programa, curso, nota, estado, creado_en
            FROM estudiantes
            ORDER BY id ASC
        """).fetchall()

    sep = request.args.get("sep", ";")
    si = StringIO()
    w = csv.writer(si, delimiter=sep, lineterminator="\n")
    w.writerow(["id", "token", "nombre", "programa", "curso", "nota", "estado", "creado_en"])

    for r in rows:
        w.writerow([
            r["id"], r["token"], r["nombre"], r["programa"], r["curso"],
            f"{r['nota']:.2f}", r["estado"], r["creado_en"]
        ])

    data = ("\ufeff" + si.getvalue()).encode("utf-8")
    buf = BytesIO(data)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name="estudiantes.csv",
        mimetype="text/csv; charset=utf-8"
    )


@app.route("/admin/export/xlsx")
@login_required
@role_required("admin")
def export_xlsx():
    with conn() as c:
        rows = c.execute("""
            SELECT id, token, nombre, programa, curso, nota, estado, creado_en
            FROM estudiantes
            ORDER BY id ASC
        """).fetchall()

    wb = Workbook()
    ws = wb.active
    ws.title = "Estudiantes"

    headers = ["id", "token", "nombre", "programa", "curso", "nota", "estado", "creado_en"]
    ws.append(headers)

    for r in rows:
        dt = r["creado_en"]
        try:
            dt = datetime.strptime(dt, "%Y-%m-%d %H:%M:%S")
        except Exception:
            pass

        ws.append([
            r["id"], r["token"], r["nombre"], r["programa"], r["curso"],
            float(r["nota"]), r["estado"], dt
        ])

    header_fill = PatternFill("solid", fgColor="DCE6F1")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="CCCCCC")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    max_row = ws.max_row
    max_col = ws.max_column

    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_all

    widths = [6, 38, 24, 14, 22, 8, 14, 22]
    for i, wdt in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = wdt

    for row in range(2, max_row + 1):
        ws.cell(row=row, column=6).number_format = "0.00"
        ws.cell(row=row, column=8).number_format = "yyyy-mm-dd hh:mm:ss"
        for col in range(1, max_col + 1):
            c = ws.cell(row=row, column=col)
            c.border = border_all
            c.alignment = Alignment(vertical="center")

    ref = f"A1:{get_column_letter(max_col)}{max_row}"
    table = Table(displayName="EstudiantesTable", ref=ref)
    style = TableStyleInfo(
        name="TableStyleMedium9",
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False
    )
    table.tableStyleInfo = style
    ws.add_table(table)
    ws.freeze_panes = "A2"

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        as_attachment=True,
        download_name="estudiantes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


# =========================================
# IMPORTAR DESDE EXCEL
# =========================================
@app.route("/admin/import/xlsx", methods=["POST"])
@login_required
@role_required("admin")
def import_xlsx():
    """
    Importa estudiantes desde .xlsx:
    - Soporta plantilla vieja: programa, nombre, etapa, nivel, nota, notas
    - Soporta plantilla nueva: programa, idioma, nivel cefr, actividades, punteo final, resultado final, notas
    """
    f = request.files.get("archivo")
    if not f or not f.filename:
        flash("Sube un archivo .xlsx.", "error")
        return redirect(url_for("admin"))

    if not f.filename.lower().endswith(".xlsx"):
        flash("El archivo debe ser Excel (.xlsx).", "error")
        return redirect(url_for("admin"))

    try:
        wb = load_workbook(filename=BytesIO(f.read()), data_only=True)
        ws = wb.active
    except Exception:
        flash("No se pudo leer el Excel. Verifica el formato.", "error")
        return redirect(url_for("admin"))

    header_map = {}
    for i in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=i).value
        key = strip_accents_py(str(v or "")).strip().lower()
        header_map[key] = i

    def col(name: str):
        return header_map.get(strip_accents_py(name).strip().lower())

    programa_col = col("programa")
    name_cols = [col("nombre completo"), col("nombre")]
    etapa_col = col("etapa")
    nivel_col = col("nivel")
    nota_col = col("nota")
    notas_col = col("notas")

    idioma_col = col("idioma")
    nivel_cefr_col = col("nivel cefr")

    cols_act = {
        "examen_1": col("examen 1"),
        "examen_2": col("examen 2"),
        "lectura": col("lectura"),
        "escritura": col("escritura"),
        "vocabulario": col("vocabulario"),
        "club_conversacion": col("club de conversacion"),
        "comprension_auditiva": col("comprension auditiva"),
        "examen_general": col("examen general"),
    }

    punteo_final_col = col("punteo final")
    resultado_final_col = col("resultado final")

    if not any(name_cols) or not etapa_col or not nivel_col:
        flash("Encabezados inválidos. Requeridos: Nombre completo (o Nombre), Etapa y Nivel.", "error")
        return redirect(url_for("admin"))

    ok, skipped = 0, 0
    now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    for r in range(2, ws.max_row + 1):
        nombre = ""
        for nc in name_cols:
            if nc:
                nombre = str(ws.cell(row=r, column=nc).value or "").strip()
                if nombre:
                    break

        programa = "Inglés"
        if programa_col:
            programa = get_programa(str(ws.cell(row=r, column=programa_col).value or "Inglés").strip())

        etapa = str(ws.cell(row=r, column=etapa_col).value or "").strip() if etapa_col else ""
        nivel = str(ws.cell(row=r, column=nivel_col).value or "").strip() if nivel_col else ""

        etapa = canonical_etapa(programa, etapa)
        nivel = canonical_nivel(programa, etapa, nivel)

        if not nombre or not etapa or not nivel:
            skipped += 1
            continue

        detalle = {}

        if idioma_col:
            detalle["idioma"] = str(ws.cell(row=r, column=idioma_col).value or "").strip()
        if nivel_cefr_col:
            detalle["nivel_cefr"] = str(ws.cell(row=r, column=nivel_cefr_col).value or "").strip()

        any_act = False
        total_act = 0.0

        for _, _, key in ACTIVIDADES:
            cidx = cols_act.get(key)
            if not cidx:
                continue

            v = ws.cell(row=r, column=cidx).value
            if v is None or str(v).strip() == "":
                continue

            any_act = True
            try:
                fv = float(v)
            except Exception:
                fv = 0.0

            detalle[key] = fv
            total_act += fv

        nota = None
        if punteo_final_col:
            vpf = ws.cell(row=r, column=punteo_final_col).value
            if vpf is not None and str(vpf).strip() != "":
                try:
                    nota = float(vpf)
                    detalle["punteo_final"] = nota
                except Exception:
                    pass

        if nota is None and any_act:
            nota = total_act
            detalle["punteo_final"] = nota

        if nota is None:
            if not nota_col:
                skipped += 1
                continue
            nota_v = ws.cell(row=r, column=nota_col).value
            try:
                nota = float(nota_v)
            except Exception:
                skipped += 1
                continue

        estado = "Aprobado" if nota >= PASSING_GRADE else "Reprobado"

        if resultado_final_col:
            rf = ws.cell(row=r, column=resultado_final_col).value
            if rf is not None and str(rf).strip() != "":
                detalle["resultado_final"] = str(rf).strip()

        curso_text = f"{etapa} - {nivel}"

        notas_url = ""
        if notas_col:
            dv = ws.cell(row=r, column=notas_col).value
            notas_url = str(dv).strip() if dv is not None else ""

        token = uuid.uuid4().hex
        with conn() as c:
            c.execute("""
                INSERT INTO estudiantes(token,nombre,curso,nota,estado,creado_en,notas,detalle_notas,programa)
                VALUES(?,?,?,?,?,?,?,?,?)
            """, (
                token, nombre, curso_text, float(nota), estado, now, notas_url,
                json.dumps(detalle, ensure_ascii=False) if detalle else None,
                programa
            ))

        build_qr(token)
        ok += 1

    flash(f"Importación finalizada. Éxitos: {ok}, Omitidos: {skipped}.", "ok")
    return redirect(url_for("admin"))


@app.route("/admin/template/xlsx")
@login_required
@role_required("admin")
def download_import_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "ImportarEstudiantes"

    headers = [
        "programa",
        "nombre completo", "etapa", "nivel",
        "idioma", "nivel cefr",
        "examen 1", "examen 2", "lectura", "escritura", "vocabulario",
        "club de conversacion", "comprension auditiva", "examen general",
        "punteo final", "resultado final",
        "notas"
    ]
    ws.append(headers)

    ws.append([
        "Inglés",
        "Juan Pérez", "Intermedia", "B1 PLUS",
        "Inglés Americano", "B1 CEFR",
        8, 9, 10, 10, 10,
        9, 8, 27,
        91, "APROBADO",
        ""
    ])

    ws.append([
        "Francés",
        "Marie Dupont", "Principiante", "A1",
        "Francés", "A1 CEFR",
        6, 7, 8, 8, 7,
        6, 6, 18,
        66, "APROBADO",
        ""
    ])

    header_fill = PatternFill("solid", fgColor="DCE6F1")
    header_font = Font(bold=True)
    header_align = Alignment(horizontal="center")
    thin = Side(style="thin", color="CCCCCC")
    border_all = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ccol in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=ccol)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = border_all

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    return send_file(
        out,
        as_attachment=True,
        download_name="plantilla_importar_estudiantes.xlsx",
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.route("/admin/update-date/<int:id>", methods=["POST"])
@login_required
@role_required("admin")
def update_fecha(id):
    creado_input = (request.form.get("creado_en") or "").strip()
    dt = _parse_datetime_local(creado_input)
    if not dt:
        flash("La fecha es inválida. Usa el selector de fecha y hora.", "error")
        return redirect(url_for("admin"))

    creado_val = dt.strftime("%Y-%m-%d %H:%M:%S")
    with conn() as c:
        c.execute("UPDATE estudiantes SET creado_en=? WHERE id=?", (creado_val, id))

    flash("Fecha actualizada.", "ok")
    return redirect(url_for("admin"))


# =========================================
# ZIP CERTIFICADOS
# =========================================
@app.route("/admin/certificados/zip", methods=["POST"])
@login_required
@role_required("admin", "subadmin")
def bulk_download_certs():
    ids = request.form.getlist("ids")
    ids = [int(x) for x in ids if str(x).isdigit()]

    if not ids:
        flash("Selecciona al menos un alumno.", "error")
        return redirect(url_for("admin"))

    if len(ids) > MAX_ZIP_ITEMS:
        flash(
            f"Demasiados alumnos seleccionados. Máximo {MAX_ZIP_ITEMS} por ZIP para evitar timeout.",
            "error"
        )
        return redirect(url_for("admin"))

    q_marks = ",".join(["?"] * len(ids))

    with conn() as c:
        rows = c.execute(
            f"SELECT id, token, nombre FROM estudiantes WHERE id IN ({q_marks})",
            ids
        ).fetchall()

    if not rows:
        flash("No se encontraron alumnos con esos IDs.", "error")
        return redirect(url_for("admin"))

    rows = sorted(rows, key=lambda r: int(r["id"]))

    tmp_dir = tempfile.gettempdir()
    zip_name = f"certificados_{datetime.now().strftime('%Y%m%d_%H%M%S')}_{uuid.uuid4().hex[:8]}.zip"
    zip_path = os.path.join(tmp_dir, zip_name)

    ok_count = 0
    fail_count = 0

    try:
        with zipfile.ZipFile(zip_path, "w", compression=zipfile.ZIP_DEFLATED) as zf:
            for r in rows:
                try:
                    pdf_bytes = _build_pdf_bytes_from_token(r["token"])
                    fname = f"{_safe_filename(r['nombre'])}_{r['id']}.pdf"
                    zf.writestr(fname, pdf_bytes)
                    ok_count += 1
                except Exception as ex:
                    fail_count += 1
                    err_name = f"ERROR_{r['id']}_{_safe_filename(r['nombre'])}.txt"
                    zf.writestr(
                        err_name,
                        f"No se pudo generar PDF para ID {r['id']} token {r['token']}.\nError: {repr(ex)}\n"
                    )

        if ok_count == 0:
            try:
                if os.path.exists(zip_path):
                    os.remove(zip_path)
            except Exception:
                pass
            flash("No se pudo generar ningún PDF. Revisa logs.", "error")
            return redirect(url_for("admin"))

        @after_this_request
        def _cleanup(response):
            try:
                if os.path.exists(zip_path):
                    os.remove(zip_path)
            except Exception:
                pass
            return response

        if fail_count > 0:
            flash(
                f"ZIP generado con advertencias: {ok_count} OK, {fail_count} fallaron (se incluyeron archivos ERROR_*.txt).",
                "error"
            )

        return send_file(
            zip_path,
            as_attachment=True,
            download_name="certificados.zip",
            mimetype="application/zip",
            conditional=True
        )

    except Exception as e:
        try:
            if os.path.exists(zip_path):
                os.remove(zip_path)
        except Exception:
            pass
        flash(f"Error generando ZIP: {repr(e)}", "error")
        return redirect(url_for("admin"))


# =========================================
# GENERADOR PDF
# =========================================
def _build_pdf_bytes_from_token(token: str) -> bytes:
    _ensure_paths()

    with conn() as cdb:
        e = cdb.execute("SELECT * FROM estudiantes WHERE token=?", (token,)).fetchone()
    if not e:
        raise ValueError(f"Token no existe: {token}")

    programa = get_programa(e["programa"] or "Inglés")

    qr_png = os.path.join(QR_DIR, f"{token}.png")
    if not os.path.exists(qr_png):
        build_qr(token)

    _register_cert_fonts()
    TITLE_FONT = _pick_font("Playfair-Bold", fallback_bold=True)
    NAME_FONT = _pick_font("GreatVibes", fallback_bold=False)
    TEXT_FONT = _pick_font("Arial", fallback_bold=False)

    buf = BytesIO()
    cpdf = canvas.Canvas(buf, pagesize=landscape(A4))
    W, H = landscape(A4)

    template_png = TEMPLATE_PNG
    if programa == "Francés" and os.path.exists(TEMPLATE_PNG_FRANCES):
        template_png = TEMPLATE_PNG_FRANCES
    elif programa == "Inglés" and os.path.exists(TEMPLATE_PNG_INGLES):
        template_png = TEMPLATE_PNG_INGLES

    try:
        if os.path.exists(template_png):
            bg = ImageReader(template_png)
            cpdf.drawImage(
                bg, 0, 0, width=W, height=H, mask="auto",
                preserveAspectRatio=True, anchor="c"
            )
        else:
            cpdf.setFillColor(colors.whitesmoke)
            cpdf.rect(0, 0, W, H, stroke=0, fill=1)
    except Exception:
        cpdf.setFillColor(colors.whitesmoke)
        cpdf.rect(0, 0, W, H, stroke=0, fill=1)

    AEA_NAVY = colors.Color(0 / 255, 47 / 255, 122 / 255)

    
    nivel_txt = (e["curso"] or "").strip()

    cpdf.setFont(TITLE_FONT, 20)
    cpdf.drawCentredString(W / 2, H - 220, nivel_txt)

    if os.path.exists(LOGO_WATERMARK):
        try:
            wm = ImageReader(LOGO_WATERMARK)
            cpdf.saveState()
            if hasattr(cpdf, "setFillAlpha"):
                cpdf.setFillAlpha(0.08)
            cpdf.drawImage(
                wm, W / 2 - 210, H / 2 - 210,
                width=420, height=420,
                mask="auto", preserveAspectRatio=True
            )
            cpdf.restoreState()
        except Exception:
            pass

    name_txt = (e["nombre"] or "").strip()
    cpdf.setFillColor(AEA_NAVY)
    cpdf.setFont(NAME_FONT, 35)
    cpdf.drawCentredString(W / 2, H / 2 + 10, name_txt)

    QR_SIZE = 90
    RIGHT_MRG = 90
    TOP_MRG = 120
    QR_X = W - RIGHT_MRG - QR_SIZE
    QR_Y = H - TOP_MRG - QR_SIZE

    try:
        qr_img = ImageReader(qr_png)
        cpdf.drawImage(
            qr_img, QR_X, QR_Y, width=QR_SIZE, height=QR_SIZE,
            mask="auto", preserveAspectRatio=True
        )
        cpdf.setFont(TEXT_FONT, 9)
        cpdf.setFillColor(AEA_NAVY)
        cpdf.drawCentredString(QR_X + QR_SIZE / 2, QR_Y - 12, "Escanea para validar")
    except Exception:
        pass

    try:
        creado = datetime.fromisoformat(e["creado_en"])
        fecha_txt = _fecha_es(creado)
    except Exception:
        fecha_txt = (e["creado_en"] or "")

    cpdf.setFillColor(AEA_NAVY)
    cpdf.setFont(TITLE_FONT, 16)
    cpdf.drawCentredString(W / 2, 60, fecha_txt)

    cpdf.showPage()
    cpdf.save()
    buf.seek(0)
    return buf.read()


# =========================================
# MAIN
# =========================================
if __name__ == "__main__":
    with app.app_context():
        init_db()
        init_users()
        seed_if_empty()
    app.run(host="0.0.0.0", port=5000, debug=True)